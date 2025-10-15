import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import glob
from openpyxl import load_workbook
from scipy import stats
from scipy.stats import f
from sklearn.metrics import cohen_kappa_score


class ResultsAnalyzer:
    """
    Analyzes validated PICO extraction results from RAG-LLM pipeline.
    
    Handles HPO (hyperparameter optimization) and consistency check results
    for multiple disease cases (NSCLC, HCC).
    Calculates recall, precision, F1 scores, and system reliability metrics.
    Includes AI vs Human comparison analysis.
    """
    
    def __init__(self, results_folder: str = "results/scored"):
        """
        Initialize the results analyzer.
        
        Args:
            results_folder: Path to folder containing scored .xlsx files
        """
        self.results_folder = Path(results_folder)
        self.cases = ["NSCLC", "HCC"]
        self.analysis_types = ["hpo", "con"]
        
        # Different scenario names for HPO vs consistency checks
        self.scenarios_hpo = ["base", "hpo1", "hpo2", "hpo3", "hpo4", "hpo5", "hpo6"]
        self.scenarios_con = ["base", "base_b", "base_c", "base_d", "base_e"]
        
        self.pico_elements = ["Population", "Comparator", "Outcome"]
        
        # Store results
        self.results = {}
        # Store precision data separately
        self.precision_data = {}
        # Store human extraction data
        self.human_data = {}
        # Store AI vs Human comparison results
        self.comparison_results = {}
        
    def read_excel_file(self, filepath: Path, case: str, scenarios: List[str]) -> Dict[str, pd.DataFrame]:
        """
        Read an Excel file and extract relevant sheets using openpyxl directly.
        
        Args:
            filepath: Path to Excel file
            case: Case name (e.g., 'NSCLC', 'HCC')
            scenarios: List of scenario names to look for
            
        Returns:
            Dictionary with scenario names as keys and DataFrames as values
        """
        # Load workbook with openpyxl directly
        wb = load_workbook(filepath, read_only=True, data_only=True)
        available_sheets = wb.sheetnames
        
        data = {}
        
        for scenario in scenarios:
            # Read PC (Population & Comparator) sheet
            pc_sheet_name = f"{case}_PC_{scenario}"
            if pc_sheet_name in available_sheets:
                ws = wb[pc_sheet_name]
                # Convert worksheet to DataFrame
                data_rows = []
                headers = None
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = row
                    else:
                        data_rows.append(row)
                df_pc = pd.DataFrame(data_rows, columns=headers)
                data[f"{scenario}_PC"] = df_pc
            
            # Read O (Outcome) sheet
            o_sheet_name = f"{case}_O_{scenario}"
            if o_sheet_name in available_sheets:
                ws = wb[o_sheet_name]
                # Convert worksheet to DataFrame
                data_rows = []
                headers = None
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = row
                    else:
                        data_rows.append(row)
                df_o = pd.DataFrame(data_rows, columns=headers)
                data[f"{scenario}_O"] = df_o
        
        wb.close()
        return data
    
    def load_precision_data(self, case: str):
        """
        Load precision data from separate precision files.
        
        Args:
            case: Case name (e.g., 'NSCLC', 'HCC')
        """
        # Construct precision filename
        precision_filename = f"{case}_precision.xlsx"
        precision_filepath = self.results_folder / precision_filename
        
        if not precision_filepath.exists():
            print(f"Warning: Precision file not found: {precision_filepath}")
            return
        
        print(f"Loading precision data for {case}...")
        print(f"  File path: {precision_filepath}")
        print(f"  File exists: {precision_filepath.exists()}")
        print(f"  File size: {precision_filepath.stat().st_size} bytes")
        
        # First, check what sheets actually exist in the file
        try:
            wb = load_workbook(precision_filepath, read_only=True, data_only=True)
            actual_sheets = wb.sheetnames
            print(f"  Actual sheets in file: {actual_sheets}")
            wb.close()
        except Exception as e:
            print(f"  ✗ Error reading workbook: {type(e).__name__}: {e}")
            return
        
        if not actual_sheets:
            print(f"  ✗ No sheets found in file. File may be corrupted or in wrong format.")
            print(f"  Try opening the file in Excel to verify it's valid.")
            return
        
        # Read only the 'base' scenario from precision file
        precision_sheets = self.read_excel_file(precision_filepath, case, ["base"])
        
        print(f"  Looking for sheet: {case}_PC_base")
        print(f"  Sheets loaded: {list(precision_sheets.keys())}")
        
        if "base_PC" in precision_sheets:
            self.precision_data[case] = precision_sheets["base_PC"]
            print(f"  ✓ Successfully loaded {len(precision_sheets['base_PC'])} rows")
            print(f"  ✓ Columns: {list(precision_sheets['base_PC'].columns)}")
            
            # Extract and preview precision scores
            test_scores = self.extract_scores(precision_sheets["base_PC"], ['Population', 'Comparator'])
            print(f"  ✓ Found {len(test_scores.get('Population', []))} Population scores")
            print(f"  ✓ Found {len(test_scores.get('Comparator', []))} Comparator scores")
            if test_scores.get('Population'):
                print(f"  ✓ Sample Population scores: {test_scores['Population'][:3]}")
            if test_scores.get('Comparator'):
                print(f"  ✓ Sample Comparator scores: {test_scores['Comparator'][:3]}")
        else:
            print(f"  ✗ Sheet 'base_PC' not found in loaded data")
            print(f"  Available keys: {list(precision_sheets.keys())}")
    
    def load_human_data(self, case: str):
        """
        Load human extraction data for comparison with AI.
        
        Args:
            case: Case name (e.g., 'NSCLC', 'HCC')
        """
        human_filename = f"{case}_human.xlsx"
        human_filepath = self.results_folder / human_filename
        
        if not human_filepath.exists():
            print(f"Warning: Human extraction file not found: {human_filepath}")
            return
        
        print(f"Loading human extraction data for {case}...")
        print(f"  File path: {human_filepath}")
        
        try:
            wb = load_workbook(human_filepath, read_only=True, data_only=True)
            actual_sheets = wb.sheetnames
            print(f"  Sheets in file: {actual_sheets}")
            wb.close()
        except Exception as e:
            print(f"  ✗ Error reading workbook: {type(e).__name__}: {e}")
            return
        
        # Read both AI (base) and human sheets
        human_sheets = self.read_excel_file(human_filepath, case, ["base", "human"])
        
        if "base_PC" in human_sheets and "human_PC" in human_sheets:
            if "base_O" in human_sheets and "human_O" in human_sheets:
                self.human_data[case] = {
                    'ai_pc': human_sheets["base_PC"],
                    'ai_o': human_sheets["base_O"],
                    'human_pc': human_sheets["human_PC"],
                    'human_o': human_sheets["human_O"]
                }
                print(f"  ✓ Successfully loaded all sheets for {case}")
            else:
                print(f"  ✗ Missing Outcome sheets")
        else:
            print(f"  ✗ Missing PC sheets")
            print(f"  Available keys: {list(human_sheets.keys())}")
    
    def extract_scores(self, df: pd.DataFrame, columns: List[str]) -> Dict[str, List[float]]:
        """
        Extract score rows from a DataFrame.
        
        Args:
            df: DataFrame containing validation scores
            columns: List of column names to extract (e.g., ['Population', 'Comparator'])
            
        Returns:
            Dictionary with column names as keys and lists of scores as values
        """
        # Filter rows ending with '_score'
        score_rows = df[df.iloc[:, 0].astype(str).str.endswith('_score', na=False)]
        
        scores = {}
        for col in columns:
            if col in score_rows.columns:
                # Extract scores, convert to float, and replace NaN/empty with 0
                col_scores = pd.to_numeric(score_rows[col], errors='coerce').fillna(0).tolist()
                scores[col] = col_scores
            else:
                scores[col] = []
        
        return scores
    
    def calculate_recall(self, scores: List[float]) -> float:
        """
        Calculate recall score (average of all scores).
        
        Args:
            scores: List of individual scores (0-1)
            
        Returns:
            Average recall score
        """
        if not scores:
            return 0.0
        return np.mean(scores)
    
    def calculate_precision(self, scores: List[float]) -> float:
        """
        Calculate precision score (average of all scores).
        
        Args:
            scores: List of individual scores (0-1)
            
        Returns:
            Average precision score
        """
        if not scores:
            return 0.0
        return np.mean(scores)
    
    def calculate_f1(self, precision: float, recall: float) -> float:
        """
        Calculate F1 score from precision and recall.
        
        Args:
            precision: Precision score
            recall: Recall score
            
        Returns:
            F1 score
        """
        if precision + recall == 0:
            return 0.0
        return 2 * (precision * recall) / (precision + recall)
    
    def calculate_coefficient_of_variation(self, values: List[float]) -> float:
        """
        Calculate coefficient of variation (CV = std/mean * 100%).
        
        Args:
            values: List of values
            
        Returns:
            Coefficient of variation as percentage
        """
        if not values or len(values) < 2:
            return 0.0
        mean_val = np.mean(values)
        if mean_val == 0:
            return 0.0
        return (np.std(values, ddof=1) / mean_val) * 100
    
    def calculate_confidence_interval(self, values: List[float], confidence: float = 0.95) -> Tuple[float, float]:
        """
        Calculate confidence interval for the mean.
        
        Args:
            values: List of values
            confidence: Confidence level (default 0.95 for 95% CI)
            
        Returns:
            Tuple of (lower_bound, upper_bound)
        """
        if not values or len(values) < 2:
            return (0.0, 0.0)
        
        n = len(values)
        mean_val = np.mean(values)
        std_val = np.std(values, ddof=1)
        se = std_val / np.sqrt(n)
        
        # Use t-distribution for small sample sizes
        t_value = stats.t.ppf((1 + confidence) / 2, n - 1)
        margin = t_value * se
        
        return (mean_val - margin, mean_val + margin)
    
    def calculate_relative_standard_error(self, values: List[float]) -> float:
        """
        Calculate relative standard error (RSE = SE/mean * 100%).
        
        Args:
            values: List of values
            
        Returns:
            Relative standard error as percentage
        """
        if not values or len(values) < 2:
            return 0.0
        
        mean_val = np.mean(values)
        if mean_val == 0:
            return 0.0
        
        std_val = np.std(values, ddof=1)
        se = std_val / np.sqrt(len(values))
        
        return (se / mean_val) * 100
    
    def calculate_iqr(self, values: List[float]) -> Tuple[float, float, float]:
        """
        Calculate inter-quartile range statistics.
        
        Args:
            values: List of values
            
        Returns:
            Tuple of (Q1, median, Q3)
        """
        if not values:
            return (0.0, 0.0, 0.0)
        
        q1 = np.percentile(values, 25)
        median = np.percentile(values, 50)
        q3 = np.percentile(values, 75)
        
        return (q1, median, q3)
    
    def calculate_icc(self, data_matrix: np.ndarray) -> Tuple[float, Tuple[float, float]]:
        """
        Calculate Intraclass Correlation Coefficient (ICC) for two-way mixed-effects model
        with absolute agreement - ICC(3,1) or ICC(A,1).
        
        This is appropriate for test-retest reliability where:
        - Rows = subjects/test cases
        - Columns = raters/runs (fixed set of raters)
        - We want absolute agreement between runs
        
        Args:
            data_matrix: 2D array where rows are subjects and columns are raters/runs
            
        Returns:
            Tuple of (ICC value, (lower 95% CI, upper 95% CI))
        """
        if data_matrix.size == 0 or data_matrix.shape[0] < 2 or data_matrix.shape[1] < 2:
            return (0.0, (0.0, 0.0))
        
        n = data_matrix.shape[0]  # number of subjects
        k = data_matrix.shape[1]  # number of raters/runs
        
        # Calculate means
        subject_means = np.mean(data_matrix, axis=1)
        rater_means = np.mean(data_matrix, axis=0)
        grand_mean = np.mean(data_matrix)
        
        # Calculate sum of squares
        # Between subjects (rows)
        SS_rows = k * np.sum((subject_means - grand_mean) ** 2)
        
        # Between raters (columns)
        SS_cols = n * np.sum((rater_means - grand_mean) ** 2)
        
        # Total sum of squares
        SS_total = np.sum((data_matrix - grand_mean) ** 2)
        
        # Residual (error)
        SS_error = SS_total - SS_rows - SS_cols
        
        # Degrees of freedom
        df_rows = n - 1
        df_cols = k - 1
        df_error = (n - 1) * (k - 1)
        
        # Mean squares
        MS_rows = SS_rows / df_rows if df_rows > 0 else 0
        MS_cols = SS_cols / df_cols if df_cols > 0 else 0
        MS_error = SS_error / df_error if df_error > 0 else 0
        
        # ICC(3,1) - Two-way mixed, absolute agreement, single measure
        # ICC = (MS_rows - MS_error) / (MS_rows + (k-1)*MS_error + k*(MS_cols - MS_error)/n)
        
        numerator = MS_rows - MS_error
        denominator = MS_rows + (k - 1) * MS_error + k * (MS_cols - MS_error) / n
        
        if denominator == 0:
            return (0.0, (0.0, 0.0))
        
        icc = numerator / denominator
        
        # Calculate 95% confidence interval using F-distribution
        # Based on Shrout & Fleiss (1979)
        alpha = 0.05
        
        F_rows = MS_rows / MS_error if MS_error > 0 else 0
        
        # Lower bound
        F_lower = F_rows / f.ppf(1 - alpha/2, df_rows, df_error)
        icc_lower = (F_lower - 1) / (F_lower + (k - 1))
        
        # Upper bound
        F_upper = F_rows / f.ppf(alpha/2, df_rows, df_error)
        icc_upper = (F_upper - 1) / (F_upper + (k - 1))
        
        # Ensure ICC is in valid range [0, 1]
        icc = max(0.0, min(1.0, icc))
        icc_lower = max(0.0, min(1.0, icc_lower))
        icc_upper = max(0.0, min(1.0, icc_upper))
        
        return (icc, (icc_lower, icc_upper))
    
    def interpret_icc(self, icc: float) -> str:
        """
        Interpret ICC value according to standard guidelines.
        
        Args:
            icc: ICC value
            
        Returns:
            Interpretation string
        """
        if icc > 0.75:
            return "Excellent"
        elif icc > 0.60:
            return "Good"
        elif icc > 0.40:
            return "Fair"
        else:
            return "Poor"
    
    def get_reliability_grade(self, cv: float) -> str:
        """
        Assign reliability grade based on coefficient of variation.
        
        Args:
            cv: Coefficient of variation (%)
            
        Returns:
            Grade string
        """
        if cv < 3:
            return "Excellent"
        elif cv < 5:
            return "Very Good"
        elif cv < 10:
            return "Good"
        elif cv < 15:
            return "Fair"
        else:
            return "Poor"
    
    def calculate_cohens_kappa(self, ai_scores: List[float], human_scores: List[float], 
                               threshold: float = 0.5) -> Tuple[float, float]:
        """
        Calculate Cohen's kappa for inter-rater agreement between AI and human.
        
        Args:
            ai_scores: List of AI scores (0-1)
            human_scores: List of human scores (0-1)
            threshold: Threshold for binary classification (default 0.5)
            
        Returns:
            Tuple of (kappa, percentage_agreement)
        """
        if not ai_scores or not human_scores or len(ai_scores) != len(human_scores):
            return (0.0, 0.0)
        
        # Convert continuous scores to binary (present/absent)
        ai_binary = [1 if score >= threshold else 0 for score in ai_scores]
        human_binary = [1 if score >= threshold else 0 for score in human_scores]
        
        # Calculate Cohen's kappa
        kappa = cohen_kappa_score(human_binary, ai_binary)
        
        # Calculate percentage agreement
        agreements = sum([1 for i in range(len(ai_binary)) if ai_binary[i] == human_binary[i]])
        pct_agreement = (agreements / len(ai_binary)) * 100
        
        return (kappa, pct_agreement)
    
    def interpret_kappa(self, kappa: float) -> str:
        """
        Interpret Cohen's kappa value according to Landis & Koch (1977).
        
        Args:
            kappa: Cohen's kappa value
            
        Returns:
            Interpretation string
        """
        if kappa > 0.80:
            return "Almost Perfect"
        elif kappa > 0.60:
            return "Substantial"
        elif kappa > 0.40:
            return "Moderate"
        elif kappa > 0.20:
            return "Fair"
        elif kappa > 0.00:
            return "Slight"
        else:
            return "Poor"
    
    def analyze_concordance(self, ai_scores: List[float], human_scores: List[float],
                           threshold: float = 0.5) -> Dict:
        """
        Analyze concordance and discordance patterns between AI and human.
        
        Args:
            ai_scores: List of AI scores (0-1)
            human_scores: List of human scores (0-1)
            threshold: Threshold for classification
            
        Returns:
            Dictionary with concordance statistics
        """
        if not ai_scores or not human_scores or len(ai_scores) != len(human_scores):
            return {}
        
        n = len(ai_scores)
        
        # Binary classifications
        ai_binary = np.array([1 if score >= threshold else 0 for score in ai_scores])
        human_binary = np.array([1 if score >= threshold else 0 for score in human_scores])
        
        # Concordance patterns
        both_correct = np.sum((ai_binary == 1) & (human_binary == 1))
        both_missed = np.sum((ai_binary == 0) & (human_binary == 0))
        ai_only = np.sum((ai_binary == 1) & (human_binary == 0))
        human_only = np.sum((ai_binary == 0) & (human_binary == 1))
        
        # Error type analysis (for discordant cases)
        discordant_indices = np.where(ai_binary != human_binary)[0]
        
        # Categorize discordance by score differences
        incomplete_extraction = []  # Small differences (0.1-0.4)
        complete_omission = []  # One has 0, other has high score (>0.6)
        over_extraction = []  # AI high, human low
        interpretation_diff = []  # Both have moderate scores (0.3-0.7)
        
        for idx in discordant_indices:
            diff = abs(ai_scores[idx] - human_scores[idx])
            ai_s = ai_scores[idx]
            human_s = human_scores[idx]
            
            if 0.1 <= diff <= 0.4 and 0.3 <= min(ai_s, human_s) <= 0.7:
                incomplete_extraction.append(idx)
            elif (ai_s == 0 and human_s > 0.6) or (human_s == 0 and ai_s > 0.6):
                complete_omission.append(idx)
            elif ai_s > 0.7 and human_s < 0.3:
                over_extraction.append(idx)
            elif 0.3 <= ai_s <= 0.7 and 0.3 <= human_s <= 0.7:
                interpretation_diff.append(idx)
        
        return {
            'n_total': n,
            'both_correct': both_correct,
            'both_correct_pct': (both_correct / n * 100) if n > 0 else 0,
            'both_missed': both_missed,
            'both_missed_pct': (both_missed / n * 100) if n > 0 else 0,
            'ai_only': ai_only,
            'ai_only_pct': (ai_only / n * 100) if n > 0 else 0,
            'human_only': human_only,
            'human_only_pct': (human_only / n * 100) if n > 0 else 0,
            'incomplete_extraction': len(incomplete_extraction),
            'incomplete_extraction_pct': (len(incomplete_extraction) / n * 100) if n > 0 else 0,
            'complete_omission': len(complete_omission),
            'complete_omission_pct': (len(complete_omission) / n * 100) if n > 0 else 0,
            'over_extraction': len(over_extraction),
            'over_extraction_pct': (len(over_extraction) / n * 100) if n > 0 else 0,
            'interpretation_diff': len(interpretation_diff),
            'interpretation_diff_pct': (len(interpretation_diff) / n * 100) if n > 0 else 0
        }
    
    def compare_ai_human_performance(self, case: str) -> Dict:
        """
        Compare AI and human extraction performance for a specific case.
        
        Args:
            case: Case name (e.g., 'NSCLC', 'HCC')
            
        Returns:
            Dictionary with comparison metrics
        """
        if case not in self.human_data:
            return {}
        
        results = {
            'case': case,
            'elements': {}
        }
        
        # Analyze Population and Comparator
        for element in ['Population', 'Comparator']:
            ai_scores = self.extract_scores(self.human_data[case]['ai_pc'], [element]).get(element, [])
            human_scores = self.extract_scores(self.human_data[case]['human_pc'], [element]).get(element, [])
            
            if ai_scores and human_scores and len(ai_scores) == len(human_scores):
                # Calculate recalls
                ai_recall = self.calculate_recall(ai_scores)
                human_recall = self.calculate_recall(human_scores)
                
                # Calculate inter-rater agreement
                kappa, pct_agreement = self.calculate_cohens_kappa(ai_scores, human_scores)
                
                # Analyze concordance
                concordance = self.analyze_concordance(ai_scores, human_scores)
                
                results['elements'][element] = {
                    'ai_recall': ai_recall,
                    'human_recall': human_recall,
                    'recall_diff': ai_recall - human_recall,
                    'kappa': kappa,
                    'kappa_interpretation': self.interpret_kappa(kappa),
                    'pct_agreement': pct_agreement,
                    'concordance': concordance,
                    'n': len(ai_scores)
                }
        
        # Analyze Outcome
        ai_scores_o = self.extract_scores(self.human_data[case]['ai_o'], ['Outcome']).get('Outcome', [])
        human_scores_o = self.extract_scores(self.human_data[case]['human_o'], ['Outcome']).get('Outcome', [])
        
        if ai_scores_o and human_scores_o and len(ai_scores_o) == len(human_scores_o):
            ai_recall = self.calculate_recall(ai_scores_o)
            human_recall = self.calculate_recall(human_scores_o)
            
            kappa, pct_agreement = self.calculate_cohens_kappa(ai_scores_o, human_scores_o)
            concordance = self.analyze_concordance(ai_scores_o, human_scores_o)
            
            results['elements']['Outcome'] = {
                'ai_recall': ai_recall,
                'human_recall': human_recall,
                'recall_diff': ai_recall - human_recall,
                'kappa': kappa,
                'kappa_interpretation': self.interpret_kappa(kappa),
                'pct_agreement': pct_agreement,
                'concordance': concordance,
                'n': len(ai_scores_o)
            }
        
        # Calculate aggregate metrics
        if results['elements']:
            all_ai_recalls = [v['ai_recall'] for v in results['elements'].values()]
            all_human_recalls = [v['human_recall'] for v in results['elements'].values()]
            all_kappas = [v['kappa'] for v in results['elements'].values()]
            all_agreements = [v['pct_agreement'] for v in results['elements'].values()]
            
            results['aggregate'] = {
                'ai_recall_mean': np.mean(all_ai_recalls),
                'human_recall_mean': np.mean(all_human_recalls),
                'recall_diff_mean': np.mean(all_ai_recalls) - np.mean(all_human_recalls),
                'kappa_mean': np.mean(all_kappas),
                'kappa_interpretation': self.interpret_kappa(np.mean(all_kappas)),
                'pct_agreement_mean': np.mean(all_agreements)
            }
        
        return results
    
    def analyze_case(self, case: str, analysis_type: str) -> Dict:
        """
        Analyze results for a specific case and analysis type.
        
        Args:
            case: Case name (e.g., 'NSCLC', 'HCC')
            analysis_type: 'hpo' or 'con'
            
        Returns:
            Dictionary containing analysis results
        """
        # Construct filename for recall data
        filename = f"{case}_{analysis_type}.xlsx"
        filepath = self.results_folder / filename
        
        if not filepath.exists():
            print(f"Warning: File not found: {filepath}")
            return {}
        
        print(f"\nAnalyzing {case} - {analysis_type.upper()}...")
        
        # Select appropriate scenario list
        scenarios = self.scenarios_hpo if analysis_type == "hpo" else self.scenarios_con
        
        # Read Excel file (recall data)
        sheets_data = self.read_excel_file(filepath, case, scenarios)
        
        # Store results for each scenario
        scenario_results = {}
        
        for scenario in scenarios:
            scenario_scores = {
                'Population': {'recall': [], 'precision': []},
                'Comparator': {'recall': [], 'precision': []},
                'Outcome': {'recall': []}
            }
            
            # Extract PC recall scores
            pc_key = f"{scenario}_PC"
            if pc_key in sheets_data:
                pc_recall_scores = self.extract_scores(
                    sheets_data[pc_key], 
                    ['Population', 'Comparator']
                )
                scenario_scores['Population']['recall'] = pc_recall_scores.get('Population', [])
                scenario_scores['Comparator']['recall'] = pc_recall_scores.get('Comparator', [])
            
            # Extract O recall scores (no precision for Outcome)
            o_key = f"{scenario}_O"
            if o_key in sheets_data:
                o_recall_scores = self.extract_scores(
                    sheets_data[o_key], 
                    ['Outcome']
                )
                scenario_scores['Outcome']['recall'] = o_recall_scores.get('Outcome', [])
            
            # Extract precision scores (only for 'base' scenario from separate file)
            if scenario == 'base' and case in self.precision_data:
                pc_precision_scores = self.extract_scores(
                    self.precision_data[case],
                    ['Population', 'Comparator']
                )
                scenario_scores['Population']['precision'] = pc_precision_scores.get('Population', [])
                scenario_scores['Comparator']['precision'] = pc_precision_scores.get('Comparator', [])
            
            # Calculate metrics for each PICO element
            scenario_results[scenario] = {}
            
            # Population
            pop_recall = self.calculate_recall(scenario_scores['Population']['recall'])
            pop_precision = self.calculate_precision(scenario_scores['Population']['precision']) if scenario == 'base' and scenario_scores['Population']['precision'] else None
            pop_f1 = self.calculate_f1(pop_precision, pop_recall) if pop_precision is not None else None
            
            scenario_results[scenario]['Population'] = {
                'recall_scores': scenario_scores['Population']['recall'],
                'recall': pop_recall,
                'precision': pop_precision,
                'f1': pop_f1,
                'n': len(scenario_scores['Population']['recall']),
                'precision_n': len(scenario_scores['Population']['precision']) if scenario_scores['Population']['precision'] else 0
            }
            
            # Comparator
            comp_recall = self.calculate_recall(scenario_scores['Comparator']['recall'])
            comp_precision = self.calculate_precision(scenario_scores['Comparator']['precision']) if scenario == 'base' and scenario_scores['Comparator']['precision'] else None
            comp_f1 = self.calculate_f1(comp_precision, comp_recall) if comp_precision is not None else None
            
            scenario_results[scenario]['Comparator'] = {
                'recall_scores': scenario_scores['Comparator']['recall'],
                'recall': comp_recall,
                'precision': comp_precision,
                'f1': comp_f1,
                'n': len(scenario_scores['Comparator']['recall']),
                'precision_n': len(scenario_scores['Comparator']['precision']) if scenario_scores['Comparator']['precision'] else 0
            }
            
            # Outcome (no precision available)
            out_recall = self.calculate_recall(scenario_scores['Outcome']['recall'])
            
            scenario_results[scenario]['Outcome'] = {
                'recall_scores': scenario_scores['Outcome']['recall'],
                'recall': out_recall,
                'precision': None,
                'f1': None,
                'n': len(scenario_scores['Outcome']['recall']),
                'precision_n': 0
            }
            
            # Calculate total metrics (average of P, C, O)
            total_recall = np.mean([pop_recall, comp_recall, out_recall])
            
            # For precision and F1, only average P and C (O doesn't have precision)
            if scenario == 'base' and pop_precision is not None and comp_precision is not None:
                total_precision = np.mean([pop_precision, comp_precision])
                total_f1 = self.calculate_f1(total_precision, total_recall)
            else:
                total_precision = None
                total_f1 = None
            
            scenario_results[scenario]['Total'] = {
                'recall': total_recall,
                'precision': total_precision,
                'f1': total_f1,
                'n': sum([
                    scenario_results[scenario]['Population']['n'],
                    scenario_results[scenario]['Comparator']['n'],
                    scenario_results[scenario]['Outcome']['n']
                ]),
                'precision_n': sum([
                    scenario_results[scenario]['Population']['precision_n'],
                    scenario_results[scenario]['Comparator']['precision_n']
                ])
            }
        
        return scenario_results
    
    def run_analysis(self):
        """
        Run analysis for all cases and analysis types.
        """
        # First, load precision data for all cases
        for case in self.cases:
            self.load_precision_data(case)
        
        # Load human extraction data
        for case in self.cases:
            self.load_human_data(case)
        
        # Then analyze recall and combine with precision
        for case in self.cases:
            for analysis_type in self.analysis_types:
                key = f"{case}_{analysis_type}"
                self.results[key] = self.analyze_case(case, analysis_type)
        
        # Perform AI vs Human comparison
        for case in self.cases:
            if case in self.human_data:
                self.comparison_results[case] = self.compare_ai_human_performance(case)
    
    def print_results(self):
        """
        Print analysis results in a clear, formatted manner.
        Order: Base scenarios first, then simulation scenarios, then reliability analysis,
        finally AI vs Human comparison.
        """
        print("\n" + "="*80)
        print("RAG-LLM PIPELINE RESULTS ANALYSIS")
        print("="*80)
        
        # ===================================================================
        # SECTION 1: BASE SCENARIO METRICS (RECALL, PRECISION, F1)
        # ===================================================================
        print("\n" + "█"*80)
        print("█" + " "*78 + "█")
        print("█" + "BASE SCENARIO RESULTS (RECALL, PRECISION, F1)".center(78) + "█")
        print("█" + " "*78 + "█")
        print("█"*80)
        
        for case in self.cases:
            # Check if precision data exists for this case
            has_precision = False
            base_metrics = None
            
            # Check both HPO and CON to find base scenario with precision
            for analysis_type in self.analysis_types:
                key = f"{case}_{analysis_type}"
                if key in self.results and 'base' in self.results[key]:
                    if self.results[key]['base']['Total']['precision'] is not None:
                        has_precision = True
                        base_metrics = self.results[key]['base']
                        break
            
            if has_precision and base_metrics:
                print(f"\n{'─'*80}")
                print(f"CASE: {case}")
                print(f"{'─'*80}")
                
                print(f"\n{'Metric':<15} {'Total':>8} {'Population':>12} {'Comparator':>12} {'Outcome':>10}")
                print("─" * 80)
                
                # Recall
                print(f"{'Recall':<15} "
                      f"{base_metrics['Total']['recall']:>8.3f} "
                      f"{base_metrics['Population']['recall']:>12.3f} "
                      f"{base_metrics['Comparator']['recall']:>12.3f} "
                      f"{base_metrics['Outcome']['recall']:>10.3f}")
                
                # N for Recall
                print(f"{'N (Recall)':<15} "
                      f"{base_metrics['Total']['n']:>8} "
                      f"{base_metrics['Population']['n']:>12} "
                      f"{base_metrics['Comparator']['n']:>12} "
                      f"{base_metrics['Outcome']['n']:>10}")
                
                print("─" * 80)
                
                # Precision
                print(f"{'Precision':<15} "
                      f"{base_metrics['Total']['precision']:>8.3f} "
                      f"{base_metrics['Population']['precision']:>12.3f} "
                      f"{base_metrics['Comparator']['precision']:>12.3f} "
                      f"{'N/A':>10}")
                
                # N for Precision
                n_pop_prec = base_metrics['Population']['precision_n']
                n_comp_prec = base_metrics['Comparator']['precision_n']
                n_total_prec = n_pop_prec + n_comp_prec
                
                print(f"{'N (Precision)':<15} "
                      f"{n_total_prec:>8} "
                      f"{n_pop_prec:>12} "
                      f"{n_comp_prec:>12} "
                      f"{'N/A':>10}")
                
                print("─" * 80)
                
                # F1
                print(f"{'F1 Score':<15} "
                      f"{base_metrics['Total']['f1']:>8.3f} "
                      f"{base_metrics['Population']['f1']:>12.3f} "
                      f"{base_metrics['Comparator']['f1']:>12.3f} "
                      f"{'N/A':>10}")
            else:
                print(f"\n{'─'*80}")
                print(f"CASE: {case}")
                print(f"{'─'*80}")
                print("No precision data available")
        
        # ===================================================================
        # SECTION 2: ALL SIMULATION SCENARIOS (RECALL ONLY)
        # ===================================================================
        print("\n\n" + "█"*80)
        print("█" + " "*78 + "█")
        print("█" + "SIMULATION SCENARIOS (RECALL COMPARISON)".center(78) + "█")
        print("█" + " "*78 + "█")
        print("█"*80)
        
        for case in self.cases:
            for analysis_type in self.analysis_types:
                key = f"{case}_{analysis_type}"
                
                if key not in self.results or not self.results[key]:
                    continue
                
                print(f"\n{'─'*80}")
                print(f"CASE: {case} | TYPE: {analysis_type.upper()}")
                print(f"{'─'*80}")
                
                results = self.results[key]
                
                # Select appropriate scenario list
                scenarios = self.scenarios_hpo if analysis_type == "hpo" else self.scenarios_con
                
                # Print RECALL results
                print(f"\n{'RECALL SCORES':^80}")
                print(f"{'Scenario':<12} {'Total':>8} {'Population':>12} {'Comparator':>12} {'Outcome':>10} {'N':>6}")
                print("─" * 80)
                
                for scenario in scenarios:
                    if scenario not in results:
                        continue
                    
                    r = results[scenario]
                    print(f"{scenario:<12} "
                          f"{r['Total']['recall']:>8.3f} "
                          f"{r['Population']['recall']:>12.3f} "
                          f"{r['Comparator']['recall']:>12.3f} "
                          f"{r['Outcome']['recall']:>10.3f} "
                          f"{r['Total']['n']:>6}")
                
                # Calculate and print summary statistics for RECALL
                print("\n" + "─" * 80)
                print("RECALL SUMMARY STATISTICS")
                print("─" * 80)
                
                # Average across scenarios
                total_recalls = [results[s]['Total']['recall'] for s in scenarios if s in results]
                pop_recalls = [results[s]['Population']['recall'] for s in scenarios if s in results]
                comp_recalls = [results[s]['Comparator']['recall'] for s in scenarios if s in results]
                out_recalls = [results[s]['Outcome']['recall'] for s in scenarios if s in results]
                
                print(f"{'Average':<12} "
                      f"{np.mean(total_recalls):>8.3f} "
                      f"{np.mean(pop_recalls):>12.3f} "
                      f"{np.mean(comp_recalls):>12.3f} "
                      f"{np.mean(out_recalls):>10.3f}")
                
                print(f"{'Std Dev':<12} "
                      f"{np.std(total_recalls):>8.3f} "
                      f"{np.std(pop_recalls):>12.3f} "
                      f"{np.std(comp_recalls):>12.3f} "
                      f"{np.std(out_recalls):>10.3f}")
                
                print(f"{'Min':<12} "
                      f"{np.min(total_recalls):>8.3f} "
                      f"{np.min(pop_recalls):>12.3f} "
                      f"{np.min(comp_recalls):>12.3f} "
                      f"{np.min(out_recalls):>10.3f}")
                
                print(f"{'Max':<12} "
                      f"{np.max(total_recalls):>8.3f} "
                      f"{np.max(pop_recalls):>12.3f} "
                      f"{np.max(comp_recalls):>12.3f} "
                      f"{np.max(out_recalls):>10.3f}")
        
        # ===================================================================
        # SECTION 3: SYSTEM RELIABILITY AND CONSISTENCY ANALYSIS
        # ===================================================================
        print("\n\n" + "█"*80)
        print("█" + " "*78 + "█")
        print("█" + "SYSTEM RELIABILITY AND CONSISTENCY ANALYSIS".center(78) + "█")
        print("█" + " "*78 + "█")
        print("█"*80)
        
        for case in self.cases:
            print(f"\n{'═'*80}")
            print(f"CASE: {case}")
            print(f"{'═'*80}")
            
            # ---------------------------------------------------------------
            # 3.1: HPO Analysis - Identify best scenario
            # ---------------------------------------------------------------
            hpo_key = f"{case}_hpo"
            if hpo_key in self.results and self.results[hpo_key]:
                print(f"\n{'▬'*80}")
                print(f"3.1 HYPERPARAMETER OPTIMIZATION ANALYSIS")
                print(f"{'▬'*80}")
                
                hpo_results = self.results[hpo_key]
                scenarios = [s for s in self.scenarios_hpo if s in hpo_results]
                
                # Find best scenario
                best_scenario = None
                best_recall = 0.0
                for scenario in scenarios:
                    recall = hpo_results[scenario]['Total']['recall']
                    if recall > best_recall:
                        best_recall = recall
                        best_scenario = scenario
                
                base_recall = hpo_results['base']['Total']['recall']
                
                print(f"\nBest Performing Scenario: {best_scenario} (Total Recall: {best_recall:.4f})")
                print(f"Baseline Scenario:        base (Total Recall: {base_recall:.4f})")
                print(f"Performance Delta:        {(best_recall - base_recall):.4f} ({((best_recall - base_recall) / base_recall * 100):+.2f}%)")
                
                # Show all HPO scenarios ranked
                print(f"\n{'All HPO Scenarios (Ranked by Total Recall):'}")
                print(f"{'─'*80}")
                print(f"{'Rank':<6} {'Scenario':<12} {'Total':>8} {'Population':>12} {'Comparator':>12} {'Outcome':>10}")
                print(f"{'─'*80}")
                
                ranked_scenarios = sorted(scenarios, 
                                        key=lambda s: hpo_results[s]['Total']['recall'], 
                                        reverse=True)
                
                for rank, scenario in enumerate(ranked_scenarios, 1):
                    r = hpo_results[scenario]
                    marker = " ★" if scenario == best_scenario else ""
                    marker += " [BASE]" if scenario == "base" else ""
                    
                    print(f"{rank:<6} {scenario:<12} "
                          f"{r['Total']['recall']:>8.4f} "
                          f"{r['Population']['recall']:>12.4f} "
                          f"{r['Comparator']['recall']:>12.4f} "
                          f"{r['Outcome']['recall']:>10.4f}{marker}")
                
                # HPO Variability Analysis
                total_recalls = [hpo_results[s]['Total']['recall'] for s in scenarios]
                hpo_cv = self.calculate_coefficient_of_variation(total_recalls)
                hpo_range = max(total_recalls) - min(total_recalls)
                
                print(f"\n{'HPO Variability:'}")
                print(f"  Coefficient of Variation: {hpo_cv:.2f}%")
                print(f"  Range (Max - Min):        {hpo_range:.4f}")
                print(f"  Relative Range:           {(hpo_range / np.mean(total_recalls) * 100):.2f}%")
            
            # ---------------------------------------------------------------
            # 3.2: Consistency Analysis - Multiple runs of base scenario
            # ---------------------------------------------------------------
            con_key = f"{case}_con"
            if con_key in self.results and self.results[con_key]:
                print(f"\n{'▬'*80}")
                print(f"3.2 CONSISTENCY ANALYSIS (BASE SCENARIO REPEATED RUNS)")
                print(f"{'▬'*80}")
                
                con_results = self.results[con_key]
                scenarios = [s for s in self.scenarios_con if s in con_results]
                n_runs = len(scenarios)
                
                print(f"\nNumber of Runs: {n_runs}")
                print(f"Scenarios: {', '.join(scenarios)}")
                
                # Collect metrics for all runs
                metrics_data = {
                    'Total': [],
                    'Population': [],
                    'Comparator': [],
                    'Outcome': []
                }
                
                for scenario in scenarios:
                    metrics_data['Total'].append(con_results[scenario]['Total']['recall'])
                    metrics_data['Population'].append(con_results[scenario]['Population']['recall'])
                    metrics_data['Comparator'].append(con_results[scenario]['Comparator']['recall'])
                    metrics_data['Outcome'].append(con_results[scenario]['Outcome']['recall'])
                
                # =========================================================
                # 1. CENTRAL TENDENCY AND DISPERSION
                # =========================================================
                print(f"\n{'─'*80}")
                print("1. CENTRAL TENDENCY AND DISPERSION")
                print("─" * 80)
                
                print(f"\n{'Metric':<15} {'Total':>12} {'Population':>12} {'Comparator':>12} {'Outcome':>10}")
                print("─" * 80)
                
                # Mean ± SD
                mean_total = np.mean(metrics_data['Total'])
                mean_pop = np.mean(metrics_data['Population'])
                mean_comp = np.mean(metrics_data['Comparator'])
                mean_out = np.mean(metrics_data['Outcome'])
                
                std_total = np.std(metrics_data['Total'], ddof=1)
                std_pop = np.std(metrics_data['Population'], ddof=1)
                std_comp = np.std(metrics_data['Comparator'], ddof=1)
                std_out = np.std(metrics_data['Outcome'], ddof=1)
                
                print(f"{'Mean ± SD':<15} "
                      f"{mean_total:>5.4f}±{std_total:<.4f} "
                      f"{mean_pop:>5.4f}±{std_pop:<.4f} "
                      f"{mean_comp:>5.4f}±{std_comp:<.4f} "
                      f"{mean_out:>5.4f}±{std_out:<.4f}")
                
                # Median
                med_total = np.median(metrics_data['Total'])
                med_pop = np.median(metrics_data['Population'])
                med_comp = np.median(metrics_data['Comparator'])
                med_out = np.median(metrics_data['Outcome'])
                
                print(f"{'Median':<15} "
                      f"{med_total:>12.4f} "
                      f"{med_pop:>12.4f} "
                      f"{med_comp:>12.4f} "
                      f"{med_out:>10.4f}")
                
                # Range (min-max)
                min_total = np.min(metrics_data['Total'])
                min_pop = np.min(metrics_data['Population'])
                min_comp = np.min(metrics_data['Comparator'])
                min_out = np.min(metrics_data['Outcome'])
                
                max_total = np.max(metrics_data['Total'])
                max_pop = np.max(metrics_data['Population'])
                max_comp = np.max(metrics_data['Comparator'])
                max_out = np.max(metrics_data['Outcome'])
                
                print(f"{'Range (Min)':<15} "
                      f"{min_total:>12.4f} "
                      f"{min_pop:>12.4f} "
                      f"{min_comp:>12.4f} "
                      f"{min_out:>10.4f}")
                
                print(f"{'Range (Max)':<15} "
                      f"{max_total:>12.4f} "
                      f"{max_pop:>12.4f} "
                      f"{max_comp:>12.4f} "
                      f"{max_out:>10.4f}")
                
                # =========================================================
                # 2. INTRACLASS CORRELATION COEFFICIENT (ICC)
                # =========================================================
                print(f"\n{'─'*80}")
                print("2. INTRACLASS CORRELATION COEFFICIENT (ICC)")
                print("   Two-way mixed-effects model, absolute agreement, single measure")
                print("─" * 80)
                
                # Prepare data matrices for ICC calculation
                # Each row is a test case, each column is a run
                icc_matrices = {}
                
                for element in ['Population', 'Comparator', 'Outcome']:
                    # Collect recall scores for each scenario
                    scores_by_scenario = []
                    for scenario in scenarios:
                        scores = con_results[scenario][element]['recall_scores']
                        scores_by_scenario.append(scores)
                    
                    # Check if all scenarios have the same number of scores
                    if len(scores_by_scenario) > 0 and len(scores_by_scenario[0]) > 0:
                        n_scores = len(scores_by_scenario[0])
                        if all(len(s) == n_scores for s in scores_by_scenario):
                            # Create matrix: rows = test cases, columns = runs
                            icc_matrices[element] = np.array(scores_by_scenario).T
                        else:
                            print(f"  Warning: Inconsistent number of scores for {element}")
                
                # Calculate ICC for Total (using all PICO elements combined)
                # Concatenate all matrices vertically
                if len(icc_matrices) > 0:
                    total_matrix = np.vstack([icc_matrices[elem] for elem in icc_matrices.keys()])
                    icc_total, ci_total_icc = self.calculate_icc(total_matrix)
                else:
                    icc_total, ci_total_icc = 0.0, (0.0, 0.0)
                
                # Calculate ICC for each element
                icc_results = {}
                for element in ['Population', 'Comparator', 'Outcome']:
                    if element in icc_matrices:
                        icc_val, ci_icc = self.calculate_icc(icc_matrices[element])
                        icc_results[element] = (icc_val, ci_icc)
                    else:
                        icc_results[element] = (0.0, (0.0, 0.0))
                
                print(f"\n{'Component':<15} {'ICC':>8} {'95% CI':>20} {'Interpretation':>15}")
                print("─" * 80)
                
                # Overall ICC
                print(f"{'Total (Overall)':<15} "
                      f"{icc_total:>8.4f} "
                      f"[{ci_total_icc[0]:>6.4f}, {ci_total_icc[1]:>6.4f}] "
                      f"{self.interpret_icc(icc_total):>15}")
                
                # Per-element ICC
                for element in ['Population', 'Comparator', 'Outcome']:
                    icc_val, ci_icc = icc_results[element]
                    print(f"{element:<15} "
                          f"{icc_val:>8.4f} "
                          f"[{ci_icc[0]:>6.4f}, {ci_icc[1]:>6.4f}] "
                          f"{self.interpret_icc(icc_val):>15}")
                
                print(f"\nInterpretation: Excellent (>0.75), Good (0.60-0.75), Fair (0.40-0.59), Poor (<0.40)")
                
                # =========================================================
                # 3. COEFFICIENT OF VARIATION (CV)
                # =========================================================
                print(f"\n{'─'*80}")
                print("3. COEFFICIENT OF VARIATION (CV)")
                print("─" * 80)
                
                cv_total = self.calculate_coefficient_of_variation(metrics_data['Total'])
                cv_pop = self.calculate_coefficient_of_variation(metrics_data['Population'])
                cv_comp = self.calculate_coefficient_of_variation(metrics_data['Comparator'])
                cv_out = self.calculate_coefficient_of_variation(metrics_data['Outcome'])
                
                print(f"\n{'Component':<15} {'CV (%)':<10} {'Interpretation':<20}")
                print("─" * 80)
                
                cv_interpretation_total = "Excellent" if cv_total < 10 else "Acceptable" if cv_total < 15 else "Poor"
                cv_interpretation_pop = "Excellent" if cv_pop < 10 else "Acceptable" if cv_pop < 15 else "Poor"
                cv_interpretation_comp = "Excellent" if cv_comp < 10 else "Acceptable" if cv_comp < 15 else "Poor"
                cv_interpretation_out = "Excellent" if cv_out < 10 else "Acceptable" if cv_out < 15 else "Poor"
                
                print(f"{'Total':<15} {cv_total:<10.2f} {cv_interpretation_total:<20}")
                print(f"{'Population':<15} {cv_pop:<10.2f} {cv_interpretation_pop:<20}")
                print(f"{'Comparator':<15} {cv_comp:<10.2f} {cv_interpretation_comp:<20}")
                print(f"{'Outcome':<15} {cv_out:<10.2f} {cv_interpretation_out:<20}")
                
                print(f"\nNote: CV < 10% is considered excellent consistency")
                
                # =========================================================
                # 4. 95% CONFIDENCE INTERVALS
                # =========================================================
                print(f"\n{'─'*80}")
                print("4. 95% CONFIDENCE INTERVALS FOR MEAN RECALL")
                print("─" * 80)
                
                ci_total = self.calculate_confidence_interval(metrics_data['Total'])
                ci_pop = self.calculate_confidence_interval(metrics_data['Population'])
                ci_comp = self.calculate_confidence_interval(metrics_data['Comparator'])
                ci_out = self.calculate_confidence_interval(metrics_data['Outcome'])
                
                print(f"\n{'Component':<15} {'Mean':>8} {'95% CI':>22} {'Width':>10}")
                print("─" * 80)
                
                print(f"{'Total':<15} "
                      f"{mean_total:>8.4f} "
                      f"[{ci_total[0]:>7.4f}, {ci_total[1]:>7.4f}] "
                      f"{ci_total[1] - ci_total[0]:>10.4f}")
                
                print(f"{'Population':<15} "
                      f"{mean_pop:>8.4f} "
                      f"[{ci_pop[0]:>7.4f}, {ci_pop[1]:>7.4f}] "
                      f"{ci_pop[1] - ci_pop[0]:>10.4f}")
                
                print(f"{'Comparator':<15} "
                      f"{mean_comp:>8.4f} "
                      f"[{ci_comp[0]:>7.4f}, {ci_comp[1]:>7.4f}] "
                      f"{ci_comp[1] - ci_comp[0]:>10.4f}")
                
                print(f"{'Outcome':<15} "
                      f"{mean_out:>8.4f} "
                      f"[{ci_out[0]:>7.4f}, {ci_out[1]:>7.4f}] "
                      f"{ci_out[1] - ci_out[0]:>10.4f}")
                
                print(f"\nNote: Narrower confidence intervals indicate more precise estimates")
                
                # =========================================================
                # 5. STABILITY METRICS
                # =========================================================
                print(f"\n{'─'*80}")
                print("5. STABILITY METRICS")
                print("─" * 80)
                
                # Absolute difference between best and worst
                range_total = max_total - min_total
                range_pop = max_pop - min_pop
                range_comp = max_comp - min_comp
                range_out = max_out - min_out
                
                # Percentage difference
                pct_diff_total = (range_total / mean_total * 100) if mean_total > 0 else 0
                pct_diff_pop = (range_pop / mean_pop * 100) if mean_pop > 0 else 0
                pct_diff_comp = (range_comp / mean_comp * 100) if mean_comp > 0 else 0
                pct_diff_out = (range_out / mean_out * 100) if mean_out > 0 else 0
                
                print(f"\n{'Component':<15} {'Abs. Diff':>12} {'% Diff':>10} {'Best Run':>12} {'Worst Run':>12}")
                print("─" * 80)
                
                # Find which runs had best and worst performance for each component
                best_run_total = scenarios[np.argmax(metrics_data['Total'])]
                worst_run_total = scenarios[np.argmin(metrics_data['Total'])]
                
                print(f"{'Total':<15} "
                      f"{range_total:>12.4f} "
                      f"{pct_diff_total:>9.2f}% "
                      f"{best_run_total:>12} "
                      f"{worst_run_total:>12}")
                
                best_run_pop = scenarios[np.argmax(metrics_data['Population'])]
                worst_run_pop = scenarios[np.argmin(metrics_data['Population'])]
                
                print(f"{'Population':<15} "
                      f"{range_pop:>12.4f} "
                      f"{pct_diff_pop:>9.2f}% "
                      f"{best_run_pop:>12} "
                      f"{worst_run_pop:>12}")
                
                best_run_comp = scenarios[np.argmax(metrics_data['Comparator'])]
                worst_run_comp = scenarios[np.argmin(metrics_data['Comparator'])]
                
                print(f"{'Comparator':<15} "
                      f"{range_comp:>12.4f} "
                      f"{pct_diff_comp:>9.2f}% "
                      f"{best_run_comp:>12} "
                      f"{worst_run_comp:>12}")
                
                best_run_out = scenarios[np.argmax(metrics_data['Outcome'])]
                worst_run_out = scenarios[np.argmin(metrics_data['Outcome'])]
                
                print(f"{'Outcome':<15} "
                      f"{range_out:>12.4f} "
                      f"{pct_diff_out:>9.2f}% "
                      f"{best_run_out:>12} "
                      f"{worst_run_out:>12}")
                
                # =========================================================
                # SUMMARY AND INTERPRETATION
                # =========================================================
                print(f"\n{'─'*80}")
                print("CONSISTENCY SUMMARY")
                print("─" * 80)
                
                # Overall assessment based on multiple criteria
                reliability_scores = []
                
                # ICC-based score (0-100)
                icc_score = icc_total * 100
                reliability_scores.append(('ICC', icc_score, self.interpret_icc(icc_total)))
                
                # CV-based score (100 - CV, capped at 0)
                cv_score = max(0, 100 - cv_total)
                cv_grade = "Excellent" if cv_total < 10 else "Acceptable" if cv_total < 15 else "Poor"
                reliability_scores.append(('CV', cv_score, cv_grade))
                
                # Stability-based score (100 - percentage difference)
                stability_score = max(0, 100 - pct_diff_total)
                stability_grade = "Excellent" if pct_diff_total < 10 else "Acceptable" if pct_diff_total < 20 else "Poor"
                reliability_scores.append(('Stability', stability_score, stability_grade))
                
                print(f"\n{'Metric':<15} {'Score':>8} {'Grade':>15}")
                print("─" * 80)
                for metric_name, score, grade in reliability_scores:
                    print(f"{metric_name:<15} {score:>8.2f} {grade:>15}")
                
                # Overall composite score
                composite_score = np.mean([s[1] for s in reliability_scores])
                if composite_score >= 75 and icc_total > 0.75 and cv_total < 10:
                    overall_grade = "EXCELLENT - System is highly reliable and deployment-ready"
                elif composite_score >= 60 and icc_total > 0.60 and cv_total < 15:
                    overall_grade = "GOOD - System shows reliable performance"
                elif composite_score >= 40:
                    overall_grade = "FAIR - System needs improvement for production use"
                else:
                    overall_grade = "POOR - System requires significant improvement"
                
                print("─" * 80)
                print(f"{'Overall':<15} {composite_score:>8.2f} {overall_grade}")
                print("─" * 80)
        
        # ===================================================================
        # SECTION 4: AI vs HUMAN COMPARISON ANALYSIS
        # ===================================================================
        self.print_human_ai_comparison()
        
        print("\n" + "="*80)
        print("ANALYSIS COMPLETE")
        print("="*80)
    
    def print_human_ai_comparison(self):
        """
        Print comprehensive AI vs Human comparison analysis.
        """
        if not self.comparison_results:
            print("\n\n" + "█"*80)
            print("█" + " "*78 + "█")
            print("█" + "AI vs HUMAN EXTRACTION COMPARISON".center(78) + "█")
            print("█" + " "*78 + "█")
            print("█"*80)
            print("\nNo human extraction data available for comparison.")
            return
        
        print("\n\n" + "█"*80)
        print("█" + " "*78 + "█")
        print("█" + "AI vs HUMAN EXTRACTION COMPARISON".center(78) + "█")
        print("█" + " "*78 + "█")
        print("█"*80)
        
        for case in self.cases:
            if case not in self.comparison_results:
                continue
            
            comp = self.comparison_results[case]
            
            if not comp or 'elements' not in comp or not comp['elements']:
                continue
            
            print(f"\n{'═'*80}")
            print(f"CASE: {case}")
            print(f"{'═'*80}")
            
            # ---------------------------------------------------------------
            # 4.1: DIRECT PERFORMANCE COMPARISON
            # ---------------------------------------------------------------
            print(f"\n{'▬'*80}")
            print(f"4.1 DIRECT PERFORMANCE COMPARISON (RECALL)")
            print(f"{'▬'*80}")
            
            print(f"\n{'Element':<15} {'AI Recall':>12} {'Human Recall':>14} {'Difference':>12} {'N':>6}")
            print("─" * 80)
            
            for element in ['Population', 'Comparator', 'Outcome']:
                if element in comp['elements']:
                    e = comp['elements'][element]
                    ai_recall = e['ai_recall']
                    human_recall = e['human_recall']
                    diff = e['recall_diff']
                    n = e['n']
                    
                    # Color coding for difference (using text indicators)
                    diff_str = f"{diff:+.4f}"
                    if diff > 0.05:
                        indicator = "↑"
                    elif diff < -0.05:
                        indicator = "↓"
                    else:
                        indicator = "≈"
                    
                    print(f"{element:<15} "
                          f"{ai_recall:>12.4f} "
                          f"{human_recall:>14.4f} "
                          f"{diff_str:>11} {indicator} "
                          f"{n:>6}")
            
            # Aggregate comparison
            if 'aggregate' in comp:
                agg = comp['aggregate']
                print("─" * 80)
                print(f"{'AGGREGATE':<15} "
                      f"{agg['ai_recall_mean']:>12.4f} "
                      f"{agg['human_recall_mean']:>14.4f} "
                      f"{agg['recall_diff_mean']:>+12.4f}")
            
            print("\nNote: ↑ = AI outperforms Human by >5%, ↓ = Human outperforms AI by >5%, ≈ = Similar performance")
            
            # ---------------------------------------------------------------
            # 4.2: INTER-RATER AGREEMENT ANALYSIS
            # ---------------------------------------------------------------
            print(f"\n{'▬'*80}")
            print(f"4.2 INTER-RATER AGREEMENT ANALYSIS")
            print(f"{'▬'*80}")
            
            print(f"\n{'Element':<15} {'Cohen Kappa':>12} {'Interpretation':>18} {'% Agreement':>14} {'N':>6}")
            print("─" * 80)
            
            for element in ['Population', 'Comparator', 'Outcome']:
                if element in comp['elements']:
                    e = comp['elements'][element]
                    kappa = e['kappa']
                    interp = e['kappa_interpretation']
                    pct_agr = e['pct_agreement']
                    n = e['n']
                    
                    print(f"{element:<15} "
                          f"{kappa:>12.4f} "
                          f"{interp:>18} "
                          f"{pct_agr:>13.2f}% "
                          f"{n:>6}")
            
            # Aggregate agreement
            if 'aggregate' in comp:
                agg = comp['aggregate']
                print("─" * 80)
                print(f"{'AGGREGATE':<15} "
                      f"{agg['kappa_mean']:>12.4f} "
                      f"{agg['kappa_interpretation']:>18} "
                      f"{agg['pct_agreement_mean']:>13.2f}%")
            
            print("\nInterpretation (Landis & Koch): Almost Perfect (>0.80), Substantial (0.60-0.80),")
            print("                               Moderate (0.40-0.60), Fair (0.20-0.40), Slight (0-0.20)")
            
            # ---------------------------------------------------------------
            # 4.3: CONCORDANCE AND DISCORDANCE PATTERNS
            # ---------------------------------------------------------------
            print(f"\n{'▬'*80}")
            print(f"4.3 CONCORDANCE AND DISCORDANCE PATTERNS")
            print(f"{'▬'*80}")
            
            for element in ['Population', 'Comparator', 'Outcome']:
                if element in comp['elements'] and 'concordance' in comp['elements'][element]:
                    e = comp['elements'][element]
                    conc = e['concordance']
                    
                    print(f"\n{element}:")
                    print("─" * 80)
                    
                    print(f"\n{'Pattern':<30} {'Count':>10} {'Percentage':>12}")
                    print("─" * 80)
                    
                    # Perfect concordance
                    print(f"{'Both Correct (Perfect)':<30} "
                          f"{conc['both_correct']:>10} "
                          f"{conc['both_correct_pct']:>11.2f}%")
                    
                    # Both missed
                    print(f"{'Both Missed (Coverage Gap)':<30} "
                          f"{conc['both_missed']:>10} "
                          f"{conc['both_missed_pct']:>11.2f}%")
                    
                    # AI only
                    print(f"{'AI Only (AI Advantage/FP)':<30} "
                          f"{conc['ai_only']:>10} "
                          f"{conc['ai_only_pct']:>11.2f}%")
                    
                    # Human only
                    print(f"{'Human Only (Needs Improvement)':<30} "
                          f"{conc['human_only']:>10} "
                          f"{conc['human_only_pct']:>11.2f}%")
                    
                    print("\n" + "─" * 80)
                    print("ERROR TYPE CATEGORIZATION (for discordant cases)")
                    print("─" * 80)
                    
                    print(f"{'Error Type':<30} {'Count':>10} {'Percentage':>12}")
                    print("─" * 80)
                    
                    # Incomplete extraction
                    print(f"{'Incomplete Extraction':<30} "
                          f"{conc['incomplete_extraction']:>10} "
                          f"{conc['incomplete_extraction_pct']:>11.2f}%")
                    
                    # Complete omission
                    print(f"{'Complete Omission':<30} "
                          f"{conc['complete_omission']:>10} "
                          f"{conc['complete_omission_pct']:>11.2f}%")
                    
                    # Over-extraction
                    print(f"{'Over-extraction (False Pos.)':<30} "
                          f"{conc['over_extraction']:>10} "
                          f"{conc['over_extraction_pct']:>11.2f}%")
                    
                    # Interpretation differences
                    print(f"{'Interpretation Differences':<30} "
                          f"{conc['interpretation_diff']:>10} "
                          f"{conc['interpretation_diff_pct']:>11.2f}%")
            
            # ---------------------------------------------------------------
            # 4.4: SUMMARY AND RECOMMENDATIONS
            # ---------------------------------------------------------------
            print(f"\n{'▬'*80}")
            print(f"4.4 SUMMARY AND RECOMMENDATIONS")
            print(f"{'▬'*80}")
            
            if 'aggregate' in comp:
                agg = comp['aggregate']
                
                # Overall assessment
                ai_recall = agg['ai_recall_mean']
                human_recall = agg['human_recall_mean']
                kappa = agg['kappa_mean']
                
                print("\nOVERALL ASSESSMENT:")
                print("─" * 80)
                
                # Performance comparison
                if abs(ai_recall - human_recall) < 0.05:
                    perf_assess = "COMPARABLE - AI and human performance are similar"
                elif ai_recall > human_recall:
                    perf_assess = f"AI ADVANTAGE - AI outperforms human by {(ai_recall - human_recall):.4f} ({((ai_recall - human_recall) / human_recall * 100):+.2f}%)"
                else:
                    perf_assess = f"HUMAN ADVANTAGE - Human outperforms AI by {(human_recall - ai_recall):.4f} ({((human_recall - ai_recall) / ai_recall * 100):+.2f}%)"
                
                print(f"Performance:  {perf_assess}")
                
                # Agreement assessment
                if kappa > 0.80:
                    agr_assess = "EXCELLENT - Almost perfect agreement between AI and human"
                elif kappa > 0.60:
                    agr_assess = "GOOD - Substantial agreement between AI and human"
                elif kappa > 0.40:
                    agr_assess = "FAIR - Moderate agreement between AI and human"
                else:
                    agr_assess = "NEEDS IMPROVEMENT - Agreement between AI and human is low"
                
                print(f"Agreement:    {agr_assess}")
                
                print("\nKEY FINDINGS:")
                print("─" * 80)
                
                # Analyze concordance patterns across all elements
                total_both_correct = sum([comp['elements'][e]['concordance']['both_correct'] 
                                        for e in ['Population', 'Comparator', 'Outcome'] 
                                        if e in comp['elements']])
                total_ai_only = sum([comp['elements'][e]['concordance']['ai_only'] 
                                   for e in ['Population', 'Comparator', 'Outcome'] 
                                   if e in comp['elements']])
                total_human_only = sum([comp['elements'][e]['concordance']['human_only'] 
                                      for e in ['Population', 'Comparator', 'Outcome'] 
                                      if e in comp['elements']])
                total_both_missed = sum([comp['elements'][e]['concordance']['both_missed'] 
                                       for e in ['Population', 'Comparator', 'Outcome'] 
                                       if e in comp['elements']])
                
                total_n = sum([comp['elements'][e]['n'] 
                             for e in ['Population', 'Comparator', 'Outcome'] 
                             if e in comp['elements']])
                
                print(f"• Perfect Concordance Rate: {(total_both_correct / total_n * 100):.2f}%")
                print(f"• AI-Only Identifications:  {(total_ai_only / total_n * 100):.2f}%")
                print(f"• Human-Only Identifications: {(total_human_only / total_n * 100):.2f}%")
                print(f"• Coverage Gaps (Both Missed): {(total_both_missed / total_n * 100):.2f}%")
                
                print("\nRECOMMENDATIONS:")
                print("─" * 80)
                
                recommendations = []
                
                # Recommendation based on agreement
                if kappa < 0.60:
                    recommendations.append("• Investigate sources of disagreement between AI and human extraction")
                    recommendations.append("• Consider additional training or calibration of extraction criteria")
                
                # Recommendation based on performance
                if ai_recall < 0.70:
                    recommendations.append("• AI recall is below 70% - system needs improvement before deployment")
                elif ai_recall < 0.80:
                    recommendations.append("• AI recall is moderate (70-80%) - consider optimization strategies")
                
                # Recommendation based on human performance
                if human_recall < 0.75:
                    recommendations.append("• Human recall indicates challenging extraction task")
                    recommendations.append("• Consider improving ground truth clarity or extraction guidelines")
                
                # Recommendation based on concordance patterns
                if total_both_missed / total_n > 0.15:
                    recommendations.append("• High coverage gap rate (>15%) indicates difficult-to-extract elements")
                    recommendations.append("• Review document quality and extraction methodology")
                
                if total_ai_only / total_n > 0.20:
                    recommendations.append("• High AI-only rate (>20%) - validate for false positives")
                
                if total_human_only / total_n > 0.20:
                    recommendations.append("• High human-only rate (>20%) - AI missing significant information")
                    recommendations.append("• Focus on improving AI recall for these cases")
                
                if not recommendations:
                    recommendations.append("• System performance is satisfactory")
                    recommendations.append("• Monitor performance on new data to ensure consistency")
                
                for rec in recommendations:
                    print(rec)
    
    def get_best_scenario(self, case: str, analysis_type: str) -> Tuple[str, float]:
        """
        Get the best performing scenario for a given case and analysis type.
        
        Args:
            case: Case name
            analysis_type: 'hpo' or 'con'
            
        Returns:
            Tuple of (scenario_name, total_recall)
        """
        key = f"{case}_{analysis_type}"
        
        if key not in self.results or not self.results[key]:
            return None, 0.0
        
        results = self.results[key]
        
        # Select appropriate scenario list
        scenarios = self.scenarios_hpo if analysis_type == "hpo" else self.scenarios_con
        
        best_scenario = None
        best_recall = 0.0
        
        for scenario in scenarios:
            if scenario not in results:
                continue
            
            recall = results[scenario]['Total']['recall']
            if recall > best_recall:
                best_recall = recall
                best_scenario = scenario
        
        return best_scenario, best_recall
    
    def get_base_metrics(self, case: str, analysis_type: str) -> Dict:
        """
        Get all metrics (recall, precision, F1) for the base scenario.
        
        Args:
            case: Case name
            analysis_type: 'hpo' or 'con'
            
        Returns:
            Dictionary with metrics for each PICO element
        """
        key = f"{case}_{analysis_type}"
        
        if key not in self.results or not self.results[key]:
            return {}
        
        if 'base' not in self.results[key]:
            return {}
        
        return self.results[key]['base']
    
    def get_consistency_metrics(self, case: str) -> Dict:
        """
        Get comprehensive consistency metrics for a case.
        
        Args:
            case: Case name
            
        Returns:
            Dictionary with consistency metrics
        """
        con_key = f"{case}_con"
        
        if con_key not in self.results or not self.results[con_key]:
            return {}
        
        con_results = self.results[con_key]
        scenarios = [s for s in self.scenarios_con if s in con_results]
        
        if not scenarios:
            return {}
        
        # Collect metrics
        total_recalls = [con_results[s]['Total']['recall'] for s in scenarios]
        
        cv = self.calculate_coefficient_of_variation(total_recalls)
        ci = self.calculate_confidence_interval(total_recalls)
        rse = self.calculate_relative_standard_error(total_recalls)
        q1, median, q3 = self.calculate_iqr(total_recalls)
        
        return {
            'n_runs': len(scenarios),
            'mean': np.mean(total_recalls),
            'std': np.std(total_recalls, ddof=1),
            'cv': cv,
            'rse': rse,
            'min': np.min(total_recalls),
            'q1': q1,
            'median': median,
            'q3': q3,
            'max': np.max(total_recalls),
            'range': np.max(total_recalls) - np.min(total_recalls),
            'iqr': q3 - q1,
            'ci_lower': ci[0],
            'ci_upper': ci[1],
            'ci_width': ci[1] - ci[0],
            'reliability_grade': self.get_reliability_grade(cv),
            'consistency_score': max(0, 100 - cv)
        }